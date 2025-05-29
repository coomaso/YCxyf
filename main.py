"""
宜昌市信用评价信息采集系统 (增强版)
版本: 4.0.0
优化内容：
1. 重构数据结构，支持资质级信用分展示
2. 智能补全缺失字段，提高数据完整性
3. 增强筛选机制，支持按资质类别精确筛选
4. 优化数据导出格式，提供更详细信用分析
"""

import logging
import sys
import os
import json
import time
import random
import base64
from dataclasses import dataclass
from typing import Dict, List, Optional, TypedDict, Any
from urllib.parse import quote
from datetime import datetime

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from Crypto.Cipher import AES

# ==================== 配置管理 ====================
@dataclass
class AppConfig:
    RETRY_COUNT: int = 3
    PAGE_SIZE: int = 10
    TIMEOUT: int = 15
    PAGE_RETRY_MAX: int = 2
    AES_KEY: bytes = os.getenv("AES_KEY", "6875616E6779696E6875616E6779696E").encode()
    AES_IV: bytes = b"sskjKingFree5138"
    EXPORT_DIR: str = "excel_output"
    LOG_FILE: str = "credit_crawler.log"
    CORE_FIELDS: set = frozenset({'cioName', 'eqtName'})  # 核心必填字段
    QUAL_DEFAULT_SCORE: float = 100.0  # 默认信用分值

    @classmethod
    def load(cls) -> 'AppConfig':
        """从环境变量加载配置"""
        return cls(
            RETRY_COUNT=int(os.getenv("RETRY_COUNT", "3")),
            PAGE_SIZE=int(os.getenv("PAGE_SIZE", "10"))
        )

# ==================== 类型定义 ====================
class QualificationRecord(TypedDict):
    """资质信用记录"""
    zzmx: str          # 资质明细
    score: float       # 诚信分值
    zxjf: float        # 专项加分
    cxdj: str          # 诚信等级
    csf: float         # 初始分
    kf: float          # 扣分
    jcf: float         # 基础分
    eqlId: str         # 资质ID

class CompanyData(TypedDict):
    """企业数据结构"""
    cioName: str        # 企业名称
    eqtName: str        # 企业资质大类
    orgId: str          # 组织ID
    cecId: str          # 信用ID
    csf: float          # 企业初始分
    qualifications: List[QualificationRecord]  # 资质信用记录列表

class SheetConfig(TypedDict):
    name: str
    filter_key: str
    filter_value: str

# ==================== 异常体系 ====================
class CrawlerError(Exception):
    """爬虫基础异常"""

class NetworkError(CrawlerError):
    """网络请求异常"""

class DecryptionError(CrawlerError):
    """数据解密异常"""

class ExportError(CrawlerError):
    """数据导出异常"""

# ==================== 日志配置 ====================
def setup_logging(config: AppConfig):
    """结构化日志配置"""
    formatter = logging.Formatter(
        "%(asctime)s - %(name)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )

    handlers = [
        logging.FileHandler(config.LOG_FILE, encoding="utf-8"),
        logging.StreamHandler()
    ]

    logging.basicConfig(
        level=logging.INFO,
        format=formatter._fmt,
        handlers=handlers
    )
    return logging.getLogger(__name__)

logger = setup_logging(AppConfig.load())

# ==================== 核心模块 ====================
class NetworkManager:
    """增强型网络请求管理器"""

    def __init__(self, config: AppConfig):
        self.config = config
        self.session = self._create_session()

    def _create_session(self) -> requests.Session:
        """创建带重试机制的会话"""
        session = requests.Session()
        retry = Retry(
            total=self.config.RETRY_COUNT,
            backoff_factor=0.3,
            status_forcelist=[500, 502, 503, 504]
        )
        adapter = HTTPAdapter(max_retries=retry)
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        return session

    def safe_request(self, url: str) -> requests.Response:
        """执行安全请求"""
        for attempt in range(1, self.config.RETRY_COUNT + 1):
            try:
                response = self.session.get(
                    url,
                    headers=self._build_headers(),
                    timeout=self.config.TIMEOUT
                )
                response.raise_for_status()
                logger.info(f"成功请求: {url}")
                return response
            except requests.RequestException as e:
                logger.warning(f"请求失败({attempt}/{self.config.RETRY_COUNT}): {str(e)}")
                if attempt == self.config.RETRY_COUNT:
                    raise NetworkError(f"请求失败: {str(e)}") from e
                time.sleep(random.uniform(1, 3))

    @staticmethod
    def _build_headers() -> Dict[str, str]:
        """构建请求头"""
        return {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/122.0.6261.95 Safari/537.36",
            "Accept": "application/json",
            "Referer": "http://106.15.60.27:22222/xxgs/"
        }

class DataExporter:
    """内存优化的数据导出器"""

    def __init__(self, config: AppConfig):
        self.config = config
        self.sheet_configs: List[SheetConfig] = [
            {"name": "全部数据", "filter_key": "", "filter_value": ""},
            {"name": "建筑工程", "filter_key": "zzmx", "filter_value": "施工总承包_建筑工程_"},
            {"name": "市政工程", "filter_key": "zzmx", "filter_value": "施工总承包_市政公用工程_"},
            {"name": "装修装饰工程", "filter_key": "zzmx", "filter_value": "专业承包_建筑装修装饰工程_"}
        ]

    def generate_report(self, data: List[CompanyData]) -> str:
        """生成多维度报告"""
        try:
            wb = Workbook(write_only=True)
            self._create_sheets(wb, data)

            filename = self._generate_filename()
            wb.save(filename)
            logger.info(f"报告生成成功: {filename}")
            return filename
        except Exception as e:
            raise ExportError(f"报告生成失败: {str(e)}") from e
        finally:
            # 确保释放资源
            if 'wb' in locals():
                wb.close()

    def _create_sheets(self, wb: Workbook, data: List[CompanyData]):
        """创建所有工作表"""
        for config in self.sheet_configs:
            sheet = wb.create_sheet(title=config["name"])
            filtered_data = self._filter_data(data, config)
            self._fill_sheet(sheet, filtered_data)

    def _filter_data(self, data: List[CompanyData], config: SheetConfig) -> List[CompanyData]:
        """基于资质的智能筛选"""
        if not config["filter_key"]:
            return data
        
        filtered = []
        for company in data:
            # 检查企业是否符合筛选条件
            matches = [
                qual for qual in company['qualifications']
                if config["filter_value"] in qual.get(config["filter_key"], "")
            ]
            
            if matches:
                # 创建企业副本并替换资质列表
                company_copy = dict(company)
                company_copy['qualifications'] = matches
                filtered.append(company_copy)
        
        return filtered

    def _fill_sheet(self, sheet, data: List[CompanyData]):
        """动态生成报表字段"""
        # 字段映射表（支持资质级和企业级字段）
        field_map = {
            'cioName': ('企业名称', 30),
            'eqtName': ('企业资质大类', 20),
            'zzmx': ('资质明细', 40),
            'score': ('诚信分值', 12),
            'csf': ('初始分', 12),
            'jcf': ('基础分', 12),
            'zxjf': ('专项加分', 12),
            'kf': ('扣分', 12),
            'cxdj': ('诚信等级', 12)
        }
        
        # 生成表头
        header = [col[0] for col in field_map.values()]
        sheet.append(header)
        
        # 填充数据（每个资质一条记录）
        for company in data:
            for qual in company['qualifications']:
                row = []
                for field, (_, _) in field_map.items():
                    # 智能选择数据源
                    if field in qual:
                        value = qual[field]
                    else:
                        value = company.get(field, '--')
                    
                    # 特殊字段处理
                    if field == 'zzmx':
                        value = str(value).replace('_', '/')
                    elif isinstance(value, float):
                        value = round(value, 2)
                    
                    row.append(value)
                sheet.append(row)
        
        # 设置列宽
        for idx, (_, width) in enumerate(field_map.values(), 1):
            sheet.column_dimensions[get_column_letter(idx)].width = width

    def _generate_filename(self) -> str:
        """生成唯一文件名"""
        os.makedirs(self.config.EXPORT_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        return os.path.join(self.config.EXPORT_DIR, f"宜昌市企业信用分析报告_{timestamp}.xlsx")

class CreditCrawler:
    """主爬虫程序"""

    def __init__(self, config: AppConfig):
        self.config = config
        self.network = NetworkManager(config)
        self.current_code: str = ""
        self.current_ts: str = ""

    def run(self) -> str:
        """执行主流程"""
        try:
            logger.info("=== 启动爬虫 ===")

            if not self._check_connectivity():
                raise NetworkError("服务器连接失败")

            self._refresh_captcha()
            total_pages = self._get_total_pages()
            data = self._crawl_pages(total_pages)

            return DataExporter(self.config).generate_report(data)
        except KeyboardInterrupt:
            logger.info("用户中断操作")
            raise
        except Exception as e:
            logger.error(f"爬虫运行失败: {str(e)}")
            raise

    def _check_connectivity(self) -> bool:
        """服务器连通性检查"""
        test_url = "http://106.15.60.27:22222"
        try:
            response = self.network.safe_request(test_url)
            return response.status_code == 200
        except NetworkError:
            return False

    def _refresh_captcha(self):
        """刷新验证码"""
        for _ in range(self.config.RETRY_COUNT):
            try:
                timestamp = str(int(time.time() * 1000))
                url = f"http://106.15.60.27:22222/ycdc/bakCmisYcOrgan/getCreateCode?codeValue={timestamp}"
                response = self.network.safe_request(url)

                result = response.json()
                if result["code"] != 0:
                    continue

                self.current_code = self._decrypt_data(result["data"])
                self.current_ts = timestamp
                logger.info("验证码刷新成功")
                return
            except Exception as e:
                logger.warning(f"验证码刷新失败: {str(e)}")

        raise NetworkError("无法获取有效验证码")

    def _decrypt_data(self, encrypted: str) -> str:
        """解密数据"""
        try:
            cipher = AES.new(self.config.AES_KEY, AES.MODE_CBC, self.config.AES_IV)
            decrypted = cipher.decrypt(base64.b64decode(encrypted))
            return decrypted.rstrip(b"\x00").decode("utf-8")
        except Exception as e:
            raise DecryptionError(f"解密失败: {str(e)}") from e

    def _get_total_pages(self) -> int:
        """获取总页数"""
        page_data = self._fetch_page(1)
        return (page_data.get("total", 0) + self.config.PAGE_SIZE - 1) // self.config.PAGE_SIZE

    def _fetch_page(self, page: int) -> Dict[str, Any]:
        """获取单页数据"""
        url = (
            "http://106.15.60.27:22222/ycdc/bakCmisYcOrgan/getCurrentIntegrityPage"
            f"?pageSize={self.config.PAGE_SIZE}&cioName=%E5%85%AC%E5%8F%B8"
            f"&page={page}&code={quote(self.current_code)}&codeValue={self.current_ts}"
        )

        response = self.network.safe_request(url)
        decrypted = self._decrypt_data(response.json().get("data", ""))

        page_data = json.loads(decrypted)
        return page_data
        
    def _crawl_pages(self, total_pages: int) -> List[CompanyData]:
        """采集所有页面数据（含智能数据校验）"""
        data = []
        
        for page in range(1, total_pages + 1):
            for attempt in range(self.config.PAGE_RETRY_MAX):
                try:
                    page_data = self._fetch_page(page)
                    page_items = page_data.get("data", [])
                    
                    # 智能数据结构转换
                    valid_items = []
                    for item in page_items:
                        # 核心字段检查
                        if not self.config.CORE_FIELDS.issubset(item.keys()):
                            logger.warning(f"跳过记录：缺失核心字段 {self.config.CORE_FIELDS - set(item.keys())}")
                            continue
                        
                        # 标准化数据结构
                        standardized = self._standardize_data(item)
                        valid_items.append(standardized)
                    
                    data.extend(valid_items)
                    logger.info(f"页码 {page}/{total_pages} | 有效数据: {len(valid_items)}")
                    break
                except Exception as e:
                    if attempt == self.config.PAGE_RETRY_MAX - 1:
                        logger.error(f"跳过第 {page} 页: {str(e)}")
                    else:
                        logger.warning(f"重试第 {page} 页: {str(e)}")
                        self._refresh_captcha()
                        time.sleep(1)
        return data

    def _standardize_data(self, raw_data: Dict) -> CompanyData:
        """标准化数据结构并补充信用分"""
        # 基础字段提取
        company = {
            'cioName': raw_data['cioName'],
            'eqtName': raw_data['eqtName'],
            'orgId': raw_data.get('orgId', ''),
            'cecId': raw_data.get('cecId', ''),
            'csf': raw_data.get('csf', self.config.QUAL_DEFAULT_SCORE),
            'qualifications': []
        }
        
        # 处理资质信用记录
        for qual in raw_data.get('zzmxcxfArray', []):
            # 资质字段智能补全
            self._supply_qual_fields(qual)
            company['qualifications'].append(qual)
        
        # 当没有资质记录时，创建默认记录
        if not company['qualifications']:
            company['qualifications'].append({
                'zzmx': company['eqtName'],
                'score': company['csf'],
                'zxjf': 0,
                'cxdj': '',
                'csf': company['csf'],
                'kf': 0,
                'jcf': 0,
                'eqlId': company['orgId'] or company['cecId'] or f"default_{time.time()}"
            })
        
        return company
    
    def _supply_qual_fields(self, qual: Dict):
        """资质字段智能补全"""
        defaults = {
            'score': qual.get('csf', self.config.QUAL_DEFAULT_SCORE),
            'zxjf': 0,
            'cxdj': '',
            'kf': 0,
            'jcf': 0
        }
        
        for field, default in defaults.items():
            if field not in qual:
                qual[field] = default
            elif isinstance(default, float) and not isinstance(qual[field], (int, float)):
                qual[field] = default

if __name__ == "__main__":
    try:
        config = AppConfig.load()
        crawler = CreditCrawler(config)
        report_path = crawler.run()
        print(f"生成报告路径: {report_path}")
        sys.exit(0)
    except CrawlerError as e:
        logger.error(f"系统错误: {str(e)}")
        sys.exit(1)
